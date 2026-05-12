#!/usr/bin/env python3
"""Build the checklist coverage matrix from the raw workbook snapshot.

The TB-030 output is a prototype-level coverage map, not final guide step
numbering. It maps each spreadsheet-tracked row to the best current route
artifact: a route block, branch prototype, option-list/default pass, appendix,
exclusion, or typed source-readiness hold.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
import re
import sys
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - handled by operator environment.
    raise SystemExit("openpyxl is required to read the raw checklist workbook.") from exc


REPO_ROOT = Path(__file__).resolve().parents[1]
RAW_WORKBOOK = REPO_ROOT / "data" / "checklist-mapping" / "raw" / "Skyrim Checklist.xlsx"
COVERAGE_MATRIX = REPO_ROOT / "data" / "checklist-mapping" / "coverage-matrix.csv"

OBJECTIVES = REPO_ROOT / "data" / "objectives" / "objectives.csv"
PROTOTYPE_MAP = REPO_ROOT / "data" / "route-planning" / "prototype-objective-block-map.csv"
LOCATION_CATALOG = REPO_ROOT / "data" / "locations" / "location-catalog.csv"
SKILL_BOOKS = REPO_ROOT / "data" / "books" / "skill-books-locations.csv"
SPELL_TOMES = REPO_ROOT / "data" / "books" / "spell-tomes-locations.csv"
BOOK_DOCUMENTS = REPO_ROOT / "data" / "books" / "book-document-locations.csv"
ENCHANTMENTS = REPO_ROOT / "data" / "skills" / "enchantment-learning-catalog.csv"
ALCHEMY = REPO_ROOT / "data" / "skills" / "alchemy-effect-catalog.csv"
PERKS = REPO_ROOT / "data" / "skills" / "perk-rank-catalog.csv"
MERCHANTS = REPO_ROOT / "data" / "skills" / "merchant-investment-catalog.csv"
ITEM_MEMBERS = REPO_ROOT / "data" / "items" / "ae-item-members.csv"
NPC_OPTIONS = REPO_ROOT / "data" / "npc" / "relationship-options.csv"

CHECKLIST_TABS = {
    "Quests",
    "Enchanting Effects",
    "Spells",
    "Dragon Shouts",
    "Locations",
    "Merchants",
    "Unique Gear",
    "Books",
    "Collectible Items",
    "Recruitable Followers",
    "Learned Alchemy Effects",
    "Perks",
}

COLUMNS = [
    "checklist_id",
    "checklist_tab",
    "checklist_entry",
    "category",
    "mapping_type",
    "guide_location",
    "branch_name",
    "exclusion_reason",
    "source_note_refs",
    "status",
    "notes",
    "objective_id",
    "matched_objective_name",
    "route_block",
    "disposition",
    "prototype_status",
    "deferred_to",
    "match_status",
    "match_source",
    "raw_sheet_row",
    "raw_cell",
    "raw_group",
    "raw_status",
    "raw_detail",
]

STATUS_VALUES = {"Y", "N", "-", "*"}
CHECKLIST_MANUAL_REVIEW_SOURCE_NOTE = "SN-000125-checklist-manual-review-reconciliation.md"


PERK_NAME_ALIASES = {
    ("pickpocket", "key master gozer"): "Keymaster",
    ("heavy armor", "reflective blows"): "Reflect Blows",
    ("two handed", "champion stance"): "Champion's Stance",
    ("two handed", "great crit char"): "Great Critical Charge",
    ("one handed", "paralysing strike"): "Paralyzing Strike",
    ("destruction", "aug flames 1"): "Augmented Flames 1",
    ("destruction", "aug flames 2"): "Augmented Flames 2",
    ("destruction", "aug frost 1"): "Augmented Frost 1",
    ("destruction", "aug frost 2"): "Augmented Frost 2",
    ("destruction", "aug shock 1"): "Augmented Shock 1",
    ("destruction", "aug shock 2"): "Augmented Shock 2",
    ("alteration", "magic res 1"): "Magic Resistance 1",
    ("alteration", "magic res 2"): "Magic Resistance 2",
    ("alteration", "magic res 3"): "Magic Resistance 3",
    ("enchanting", "insight enchanter"): "Insightful Enchanter",
    ("illusion", "master mind"): "Master of the Mind",
}

QUEST_PERK_OBJECTIVES = {
    "sinderion s serendipity": "OBJ-000810",
    "sailor s repose": "OBJ-000809",
    "prowler s profit": "OBJ-000808",
    "agent of mara": "OBJ-000803",
    "agent of dibella": "OBJ-000802",
    "nightingale armor": "OBJ-001766",
    "ancient knowledge": "OBJ-000804",
    "dragon infusion": "OBJ-000805",
    "turn of the seasons": "OBJ-000813",
}

TRANSFORMATION_PERK_OBJECTIVES = {
    "vampire lord perks": "OBJ-000817",
    "werewolf perks": "OBJ-000815",
}

QUEST_ENTRY_OVERRIDES = {
    "visit the shrine of azura": "OBJ-000165",
    "speak to the witchhunter": "OBJ-000174",
    "investigate the boethiah cultist": "OBJ-000166",
    "investigate the hall of the dead": "OBJ-000175",
    "speak to lod": "OBJ-000167",
    "darkness returns thieves guild": "OBJ-000046",
    "find kesh at the peryite shrine": "OBJ-000176",
    "drink with sam": "OBJ-000177",
    "visit the museum in dawnstar": "OBJ-000171",
    "investigate dervenin": "OBJ-000178",
    "ask about balgruuf s strange children": "OBJ-000172",
    "lair s retreat": "OBJ-002758",
    "recover andurs amulet of arkay": "OBJ-000323",
    "read eltrys note": "OBJ-001135",
    "kolbjorn barrow misc objectives": "OBJ-000465",
    "sell stalhrim equipment": "OBJ-000469",
    "thirsk rumors": "OBJ-000455",
    "nordic jewlery": "OBJ-000630",
    "noric jewlery": "OBJ-000630",
    "the bonds of matrimony": "OBJ-001945",
    "dragon research": "OBJ-000805",
    "scare my enemy": "OBJ-000104",
}

LITANY_OF_LARCENY_ENTRIES = {
    "bust of the gray fox",
    "dwemer puzzle cube",
    "east empire shipping map",
    "honningbrew decanter",
    "left eye of the falmer",
    "model ship",
    "queen bee statue",
}

CITY_HOME_BY_GROUP = {
    "windhelm": "OBJ-001923",
    "solitude": "OBJ-001924",
    "markarth": "OBJ-001922",
    "riften": "OBJ-001921",
    "whiterun": "OBJ-001920",
}

HEARTHFIRE_HOME_BY_GROUP = {
    "falkreath": "OBJ-000395",
    "morthal": "OBJ-000396",
    "dawnstar": "OBJ-000397",
}

REPRESENTATIVE_ACTIVITY_OBJECTIVES = {
    "chop wood": "OBJ-002762",
    "gather wheat": "OBJ-002763",
    "mine ore": "OBJ-002764",
    "fight fight": "OBJ-002765",
    "quest all beggars have": "OBJ-002766",
    "quest all drunks have": "OBJ-002767",
}

BOOK_TITLE_OVERRIDES = {
    "death blow of ebernanit": "OBJ-000836",
    "captains journal": "OBJ-001374",
    "everthra s journal": "OBJ-001392",
    "guard dossier antonius": "OBJ-001411",
    "guard dossier bjormund wind strider": "OBJ-001412",
    "guard dossier yakhtu gra orkulg": "OBJ-001413",
    "necromancer s journal plague of dead": "OBJ-001468",
}

UNIQUE_ITEM_OVERRIDES = {
    "tumblerbane glove": "OBJ-001700",
    "ancient nordic pickaxe": "OBJ-000425",
}

SPELL_OVERRIDES = {
    "summon unbound dremora": "OBJ-000130",
    "summon arniel s shade": "OBJ-000117",
    "arniel s convection": "OBJ-000117",
    "vision of the tenth eye": "OBJ-000132",
}

ENCHANTMENT_ENTRY_ALIASES = {
    # Raw checklist wording reverses the source-listed Enchanting Effects name.
    "damage stamina": "stamina damage",
}

ALCHEMY_INGREDIENT_ALIASES = {
    # Raw checklist spelling; UESP and the source catalog use Kresh Fiber.
    "kesh fiber": "kresh fiber",
}

PET_TELEPORT_OBJECTIVES = {
    "arachnia": "OBJ-000675",
    "bone wolf": "OBJ-000671",
    "demented elytra nymph": "OBJ-000677",
    "hilda": "OBJ-000672",
    "manic elytra nymph": "OBJ-000678",
    "skritch": "OBJ-000676",
    "sweet roll": "OBJ-000673",
    "thistle": "OBJ-000674",
}

COLLECTIBLE_ENTRY_OVERRIDES: dict[str, str] = {}

SOURCE_READINESS_BY_CATEGORY = {
    "alchemy_effect": (
        "TB-031E/TB-031H alchemy source-readiness review",
        "checklist_alchemy_source_readiness",
        "Checklist-only alchemy row; TB-031E/TB-031H must validate source scope before route inclusion.",
    ),
    "book_document": (
        "TB-031H/TB-036 book/document source-readiness review",
        "checklist_book_document_source_readiness",
        "Checklist-only book/document row; TB-031H must validate source/objective-table readiness before final appendix or route inclusion.",
    ),
    "collectible_item": (
        "TB-031F/TB-031H collectible source-readiness review",
        "checklist_collectible_source_readiness",
        "Checklist collectible row lacks a current source-backed objective match; TB-031F/TB-031H must validate or exclude it explicitly.",
    ),
    "enchantment": (
        "TB-031E/TB-031H enchantment source-readiness review",
        "checklist_enchantment_source_readiness",
        "Checklist enchantment row lacks a current source-backed learning objective; TB-031E/TB-031H must validate route scope.",
    ),
    "location": (
        "TB-031G/TB-031H location source-readiness review",
        "checklist_location_source_readiness",
        "Checklist location row lacks a current source-backed location-catalog objective; TB-031G/TB-031H must validate scope before route use.",
    ),
    "quest": (
        "TB-031F/TB-031H quest source-readiness review",
        "checklist_quest_source_readiness",
        "Checklist quest/sub-objective row lacks a current source-backed objective match; TB-031F/TB-031H must promote, map, or exclude it after TB-031C escalation review.",
    ),
    "skill_book": (
        "TB-031H/TB-036 skill-book source-readiness review",
        "checklist_skill_book_source_readiness",
        "Checklist skill-book row lacks a current title-table match; TB-031H must validate source spelling/scope before route inclusion.",
    ),
    "spell": (
        "TB-031E/TB-031H spell source-readiness review",
        "checklist_spell_source_readiness",
        "Checklist spell row lacks a current source-backed spell or parent objective match; TB-031E/TB-031H must validate route scope.",
    ),
    "unique_item": (
        "TB-031H/TB-036 unique-gear source-readiness review",
        "checklist_unique_gear_source_readiness",
        "Checklist unique-gear row lacks a current source-backed objective match; TB-031H must validate source/objective readiness before final inclusion.",
    ),
}


@dataclass(frozen=True)
class ChecklistEntry:
    tab: str
    row: int
    cell: str
    entry: str
    category: str
    group: str = ""
    raw_status: str = ""
    detail: str = ""
    key: str = ""


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_formula_string(text: str, start: int) -> tuple[str, int] | None:
    if start >= len(text) or text[start] != '"':
        return None
    out: list[str] = []
    index = start + 1
    while index < len(text):
        char = text[index]
        if char == '"':
            if index + 1 < len(text) and text[index + 1] == '"':
                out.append('"')
                index += 2
                continue
            return "".join(out), index + 1
        out.append(char)
        index += 1
    return None


def hyperlink_label(value: Any) -> str:
    text = clean_text(value)
    if not text.startswith("="):
        return text
    lowered = text.lower()
    if not lowered.startswith("=hyperlink("):
        return ""
    index = text.find("(") + 1
    first = parse_formula_string(text, index)
    if not first:
        return ""
    index = first[1]
    while index < len(text) and text[index] in " ,":
        index += 1
    second = parse_formula_string(text, index)
    if not second:
        return ""
    return clean_text(second[0])


def cell_text(cell: Any) -> str:
    if getattr(cell, "hyperlink", None) is not None and cell.value:
        return clean_text(cell.value)
    return hyperlink_label(cell.value)


def normalize(text: str) -> str:
    text = clean_text(text)
    text = text.replace("'", "'")
    text = re.sub(
        r"\((ae|cc|db|dg|hf|dawnguard|dragonborn|hearthfire|anniversary edition|creation club|optional|two versions|dragonborn only)\)",
        "",
        text,
        flags=re.IGNORECASE,
    )
    stripped_suffix = re.sub(
        r"\b(ae|cc|db|dg|hf|dawnguard|dragonborn|hearthfire|optional|two versions|dragonborn only)\b$",
        "",
        text,
        flags=re.IGNORECASE,
    ).strip()
    if stripped_suffix:
        text = stripped_suffix
    text = re.sub(r"^spell tome:\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^(artifact|unique item|leveled reward|skill book):\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^learn enchantment:\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^discover alchemy effects:\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^(clear|discover)\s+", "", text, flags=re.IGNORECASE)
    text = text.replace("&", "and")
    text = re.sub(r"[^a-z0-9]+", " ", text.lower())
    return re.sub(r"\s+", " ", text).strip()


def add_alias(index: dict[str, list[str]], alias: str, objective_id: str) -> None:
    key = normalize(alias)
    if not key:
        return
    index.setdefault(key, [])
    if objective_id not in index[key]:
        index[key].append(objective_id)
    if key.startswith("the "):
        without_article = key.removeprefix("the ")
        index.setdefault(without_article, [])
        if objective_id not in index[without_article]:
            index[without_article].append(objective_id)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def load_indexes() -> dict[str, Any]:
    objective_rows = {row["objective_id"]: row for row in read_csv(OBJECTIVES)}
    prototype_rows = {row["objective_id"]: row for row in read_csv(PROTOTYPE_MAP)}

    aliases: dict[str, list[str]] = {}
    for row in objective_rows.values():
        objective_id = row["objective_id"]
        name = row["objective_name"]
        add_alias(aliases, name, objective_id)
        add_alias(aliases, re.sub(r"\s*\([^)]*\)", "", name), objective_id)
        for prefix in ("Artifact: ", "Unique Item: ", "Leveled Reward: ", "Skill Book: ", "Spell Tome: "):
            if name.startswith(prefix):
                add_alias(aliases, name.removeprefix(prefix), objective_id)
                add_alias(aliases, re.sub(r"\s*\([^)]*\)", "", name.removeprefix(prefix)), objective_id)
        if name.startswith("Learn Enchantment: "):
            add_alias(aliases, name.removeprefix("Learn Enchantment: "), objective_id)
        if name.startswith("Discover Alchemy Effects: "):
            add_alias(aliases, name.removeprefix("Discover Alchemy Effects: "), objective_id)
        for reward in row.get("unique_rewards", "").split("|"):
            add_alias(aliases, reward, objective_id)
            add_alias(aliases, re.sub(r"\s+(handled|delivered|deferred)$", "", reward.strip(), flags=re.IGNORECASE), objective_id)
            add_alias(
                aliases,
                re.sub(r"\s+spell$", "", re.sub(r"^learn\s+", "", reward.strip(), flags=re.IGNORECASE), flags=re.IGNORECASE),
                objective_id,
            )
        for checklist_name in row.get("checklist_mapping", "").split("|"):
            add_alias(aliases, checklist_name, objective_id)
            add_alias(aliases, re.sub(r"\s+(handled|delivered|deferred)$", "", checklist_name.strip(), flags=re.IGNORECASE), objective_id)
            add_alias(
                aliases,
                re.sub(r"\s+spell$", "", re.sub(r"^learn\s+", "", checklist_name.strip(), flags=re.IGNORECASE), flags=re.IGNORECASE),
                objective_id,
            )

    book_titles: dict[str, str] = {}
    for path in (SKILL_BOOKS, SPELL_TOMES, BOOK_DOCUMENTS):
        for row in read_csv(path):
            book_titles[normalize(row["book_title"])] = row["objective_id"]
            add_alias(aliases, row["book_title"], row["objective_id"])

    locations: dict[str, str] = {}
    for row in read_csv(LOCATION_CATALOG):
        location_name = row["location_name"]
        objective_id = row["objective_id"]
        for alias in {
            location_name,
            re.sub(r"\s*\([^)]*\)", "", location_name),
            location_name.replace("Caverns", "Cavern"),
            location_name.replace("Excavation", "").strip(),
            f"{location_name} Camp",
        }:
            locations[normalize(alias)] = objective_id
        add_alias(aliases, row["location_name"], row["objective_id"])

    enchantments = {normalize(row["enchantment_name"]): row["objective_id"] for row in read_csv(ENCHANTMENTS)}
    for row in read_csv(ENCHANTMENTS):
        add_alias(aliases, row["enchantment_name"], row["objective_id"])

    alchemy: dict[tuple[str, str], str] = {}
    for row in read_csv(ALCHEMY):
        ingredient = row["ingredient_name"]
        for effect_field in ("effect_1", "effect_2", "effect_3", "effect_4"):
            effect = row[effect_field]
            alchemy[(normalize(ingredient), normalize(effect))] = row["objective_id"]
            alchemy[(normalize(ingredient), normalize(re.sub(r"\s*\([^)]*\)", "", effect)))] = row["objective_id"]
        add_alias(aliases, ingredient, row["objective_id"])

    perks: dict[tuple[str, str], str] = {}
    for row in read_csv(PERKS):
        skill = row["skill_name"]
        node = row["perk_node_name"]
        rank = row["perk_rank"]
        objective_id = row["parent_perk_tree_objective_id"]
        candidates = {
            node,
            f"{node} {rank}",
            f"{node} Rank {rank}",
            re.sub(rf"\b{re.escape(skill)}\b", "", node, flags=re.IGNORECASE).strip(),
            f"{re.sub(rf'\\b{re.escape(skill)}\\b', '', node, flags=re.IGNORECASE).strip()} {rank}".strip(),
        }
        if skill == "Lockpicking" and node.endswith(" Locks"):
            candidates.add(node.removesuffix(" Locks"))
        if node.endswith("Smithing"):
            candidates.add(node.removesuffix("ing"))
        if node == "Arcane Blacksmith":
            candidates.add("Arcane Smith")
        if node == "Advanced Armors":
            candidates.add("Adv. Armors")
        if node == "Wind Walker":
            candidates.add("Windwalker")
        if node == "Snakeblood":
            candidates.add("Snake Blood")
        if node == "Muffled Movement":
            candidates.add("Muffled Moves")
        if node == "Elemental Protection":
            candidates.add("Element Protect")
        if node == "Tower of Strength":
            candidates.add("Tower of Str.")
        for candidate in candidates:
            perks[(normalize(skill), normalize(candidate))] = objective_id

    # Known display abbreviation from the raw checklist.
    perks[(normalize("Alchemy"), normalize("ConcenPoison"))] = perks.get(
        (normalize("Alchemy"), normalize("Concentrated Poison")),
        "",
    )

    merchants: dict[str, str] = {}
    merchant_excluded: dict[str, str] = {}
    for row in read_csv(MERCHANTS):
        merchant_name = row["merchant_name"]
        for alias in re.split(r";|\(|\)", merchant_name):
            alias = alias.strip()
            if alias:
                if row["objective_id"]:
                    merchants[normalize(alias)] = row["objective_id"]
                else:
                    merchant_excluded[normalize(alias)] = row["notes"]
        if row["objective_id"]:
            merchants[normalize(merchant_name)] = row["objective_id"]
        else:
            merchant_excluded[normalize(merchant_name)] = row["notes"]
    item_members: dict[str, str] = {}
    for row in read_csv(ITEM_MEMBERS):
        item_members[normalize(row["item_name"])] = row["existing_objective_id"] or row["parent_objective_id"]
        add_alias(aliases, row["item_name"], row["existing_objective_id"] or row["parent_objective_id"])

    npc_options: dict[str, list[str]] = {}
    for row in read_csv(NPC_OPTIONS):
        names = [row["name"]]
        names.extend(re.split(r",|\bor\b", row["name"]))
        for name in names:
            key = normalize(name)
            if not key:
                continue
            npc_options.setdefault(key, []).append(row["option_id"])

    return {
        "objective_rows": objective_rows,
        "prototype_rows": prototype_rows,
        "aliases": aliases,
        "book_titles": book_titles,
        "locations": locations,
        "enchantments": enchantments,
        "alchemy": alchemy,
        "perks": perks,
        "merchants": merchants,
        "merchant_excluded": merchant_excluded,
        "item_members": item_members,
        "npc_options": npc_options,
    }


def status_cell(ws: Any, row: int, column: int) -> str:
    return clean_text(ws.cell(row=row, column=column).value).upper()


def parse_workbook(path: Path) -> list[ChecklistEntry]:
    workbook = openpyxl.load_workbook(path, data_only=False)
    entries: list[ChecklistEntry] = []

    def add(tab: str, row: int, column: int, entry: str, category: str, group: str = "", status: str = "", detail: str = "", key: str = "") -> None:
        entry = clean_text(entry)
        if not entry or entry in {"-", "HIDDEN"}:
            return
        entries.append(
            ChecklistEntry(
                tab=tab,
                row=row,
                cell=f"{openpyxl.utils.get_column_letter(column)}{row}",
                entry=entry,
                category=category,
                group=clean_text(group),
                raw_status=clean_text(status),
                detail=clean_text(detail),
                key=key,
            )
        )

    ws = workbook["Quests"]
    groups = {1: "", 6: ""}
    for row in range(4, ws.max_row + 1):
        for start_col in (1, 6):
            name_col = start_col + 1
            status = status_cell(ws, row, start_col)
            name = cell_text(ws.cell(row=row, column=name_col))
            group_text = clean_text(ws.cell(row=row, column=start_col).value)
            if status not in {"Y", "N"} and group_text and not name and not group_text.startswith("="):
                groups[start_col] = group_text
                continue
            if name and status in {"Y", "N"}:
                desc = cell_text(ws.cell(row=row, column=name_col + 1)) or cell_text(ws.cell(row=row, column=name_col + 2))
                giver = cell_text(ws.cell(row=row, column=name_col + 2)) or cell_text(ws.cell(row=row, column=9))
                detail = " | ".join(
                    part
                    for part in [
                        desc,
                        f"Giver: {giver}" if giver and giver != desc else "",
                    ]
                    if part
                )
                add("Quests", row, name_col, name, "quest", groups[start_col], status, detail)

    ws = workbook["Enchanting Effects"]
    for row in range(3, ws.max_row + 1):
        status = status_cell(ws, row, 2)
        if status in {"Y", "N"}:
            add("Enchanting Effects", row, 1, cell_text(ws.cell(row=row, column=1)), "enchantment", "", status, cell_text(ws.cell(row=row, column=14)))

    ws = workbook["Spells"]
    for row in range(3, ws.max_row + 1):
        status = status_cell(ws, row, 2)
        if status in {"Y", "N"}:
            detail = " | ".join(
                part
                for part in [
                    f"School: {cell_text(ws.cell(row=row, column=3))}",
                    f"Level: {cell_text(ws.cell(row=row, column=4))}",
                    f"Source: {cell_text(ws.cell(row=row, column=7))}",
                ]
                if part.split(": ", 1)[-1]
            )
            add("Spells", row, 1, cell_text(ws.cell(row=row, column=1)), "spell", "", status, detail)

    ws = workbook["Dragon Shouts"]
    current_shout = ""
    for row in range(3, ws.max_row + 1):
        status = status_cell(ws, row, 1)
        word = cell_text(ws.cell(row=row, column=2))
        if status not in {"Y", "N"} and word:
            current_shout = word
            continue
        if status in {"Y", "N"} and word:
            detail = " | ".join(
                part
                for part in [
                    f"Translation: {cell_text(ws.cell(row=row, column=3))}",
                    f"Word wall: {cell_text(ws.cell(row=row, column=4))}",
                ]
                if part.split(": ", 1)[-1]
            )
            add("Dragon Shouts", row, 2, f"{current_shout}: {word}", "dragon_shout_word", current_shout, status, detail, key=current_shout)

    ws = workbook["Locations"]
    for row in range(6, ws.max_row + 1):
        name = cell_text(ws.cell(row=row, column=1))
        if not name or name == "-":
            continue
        states = {
            "discovered": status_cell(ws, row, 4),
            "visited": status_cell(ws, row, 5),
            "cleared": status_cell(ws, row, 6),
        }
        if not any(value in STATUS_VALUES for value in states.values()):
            continue
        detail = f"Type: {cell_text(ws.cell(row=row, column=2))} | Region: {cell_text(ws.cell(row=row, column=3))} | States: {states}"
        add("Locations", row, 1, name, "location", cell_text(ws.cell(row=row, column=2)), "/".join(states.values()), detail)

    ws = workbook["Merchants"]
    for row in range(4, ws.max_row + 1):
        name = cell_text(ws.cell(row=row, column=1))
        if not name:
            continue
        investable = cell_text(ws.cell(row=row, column=10))
        category = "merchant_investment" if investable.lower().startswith("yes") else "merchant_reference"
        detail = " | ".join(
            part
            for part in [
                f"Partner: {cell_text(ws.cell(row=row, column=2))}" if cell_text(ws.cell(row=row, column=2)) else "",
                f"Region: {cell_text(ws.cell(row=row, column=6))}",
                f"Store: {cell_text(ws.cell(row=row, column=7))}",
                f"Merchandise: {cell_text(ws.cell(row=row, column=8))}",
                f"Investable: {investable}",
                f"Master Trader: {cell_text(ws.cell(row=row, column=11))}",
            ]
            if part.split(": ", 1)[-1]
        )
        add("Merchants", row, 1, name, category, cell_text(ws.cell(row=row, column=6)), status_cell(ws, row, 3), detail)

    ws = workbook["Unique Gear"]
    group = ""
    for row in range(3, ws.max_row + 1):
        status = status_cell(ws, row, 4)
        name = cell_text(ws.cell(row=row, column=1))
        if status not in {"Y", "N"} and name and "Gear" in name:
            group = name
            continue
        if status in {"Y", "N"}:
            detail = " | ".join(
                part
                for part in [
                    f"Type: {cell_text(ws.cell(row=row, column=3))}",
                    cell_text(ws.cell(row=row, column=5)),
                ]
                if part
            )
            add("Unique Gear", row, 1, name, "unique_item", group, status, detail)

    ws = workbook["Books"]
    group = "Regular books, journals and spell tomes"
    for row in range(4, ws.max_row + 1):
        name = cell_text(ws.cell(row=row, column=1))
        status = status_cell(ws, row, 2)
        read_status = status_cell(ws, row, 3)
        shelf_status = status_cell(ws, row, 4)
        if name == "Books to turn in for a quest or unique spell tomes":
            group = name
            continue
        if not name or name.startswith("Total:"):
            continue
        if status not in STATUS_VALUES and read_status not in STATUS_VALUES and shelf_status not in STATUS_VALUES:
            continue
        book_type = cell_text(ws.cell(row=row, column=18))
        detail = " | ".join(
            part
            for part in [
                f"Author: {cell_text(ws.cell(row=row, column=15))}" if cell_text(ws.cell(row=row, column=15)) else "",
                f"Type: {book_type}",
                f"Description: {cell_text(ws.cell(row=row, column=16))}" if cell_text(ws.cell(row=row, column=16)) else "",
            ]
            if part
        )
        category = "book_document"
        if book_type.lower().startswith("spell tome"):
            category = "spell_tome_book"
        elif book_type.lower().startswith("skill book"):
            category = "skill_book"
        elif name.lower().startswith("black book:"):
            category = "black_book"
        elif "regular books" in group.lower() and book_type.lower() in {"book", "journal", "note"}:
            category = "general_book"
        add("Books", row, 1, name, category, group, f"{status}/{read_status}/{shelf_status}", detail)

    ws = workbook["Collectible Items"]
    blocks = [
        (1, "Bugs in a Jar"),
        (4, "Dragon Claws"),
        (7, "Treasure Maps"),
        (10, "Miscellaneous Collectibles"),
        (13, "Paragons"),
    ]
    for start_col, group in blocks:
        for row in range(4, ws.max_row + 1):
            status = status_cell(ws, row, start_col)
            name = cell_text(ws.cell(row=row, column=start_col + 1))
            if status in {"Y", "N"} and name:
                detail = cell_text(ws.cell(row=row, column=start_col + 2))
                add("Collectible Items", row, start_col + 1, name, "collectible_item", group, status, detail)

    ws = workbook["Recruitable Followers"]
    for row in range(3, ws.max_row + 1):
        status = status_cell(ws, row, 2)
        name = cell_text(ws.cell(row=row, column=1))
        if status in {"Y", "N"} and name:
            flags = [
                f"Alive: {status_cell(ws, row, 3)}",
                f"Prerequisite: {cell_text(ws.cell(row=row, column=5))}",
                f"Marry: {cell_text(ws.cell(row=row, column=8))}",
                f"Blades: {cell_text(ws.cell(row=row, column=9))}",
                f"Steward: {cell_text(ws.cell(row=row, column=10))}",
                f"Pet: {cell_text(ws.cell(row=row, column=11))}",
                f"Trainer: {cell_text(ws.cell(row=row, column=12))}",
            ]
            add("Recruitable Followers", row, 1, name, "follower_option", "", status, " | ".join(flags))

    ws = workbook["Learned Alchemy Effects"]
    for row in range(3, ws.max_row + 1):
        status = status_cell(ws, row, 1)
        ingredient = cell_text(ws.cell(row=row, column=2))
        effect = cell_text(ws.cell(row=row, column=3))
        if status in {"Y", "N"} and ingredient and effect:
            detail = " | ".join(
                part
                for part in [
                    f"Ingredient: {ingredient}",
                    f"Effect: {effect}",
                    f"Value: {cell_text(ws.cell(row=row, column=7))}",
                    f"Garden: {cell_text(ws.cell(row=row, column=8))}",
                ]
                if part.split(": ", 1)[-1]
            )
            add("Learned Alchemy Effects", row, 2, f"{ingredient}: {effect}", "alchemy_effect", ingredient, status, detail, key=f"{ingredient}|{effect}")

    ws = workbook["Perks"]
    current_skill: dict[int, str] = {1: "", 5: "", 9: ""}
    for row in range(1, ws.max_row + 1):
        for start_col in (1, 5, 9):
            maybe_skill = cell_text(ws.cell(row=row, column=start_col))
            obtained = status_cell(ws, row, start_col + 1)
            perk = cell_text(ws.cell(row=row, column=start_col + 2))
            if maybe_skill and obtained not in {"Y", "N"} and not perk and not maybe_skill.lower().startswith(("req", "-", "stealth", "warrior", "mage")):
                current_skill[start_col] = maybe_skill
                continue
            if obtained in {"Y", "N"} and perk and current_skill[start_col]:
                detail = f"Req: {maybe_skill} | Prereq: {cell_text(ws.cell(row=row, column=start_col + 3))}"
                add("Perks", row, start_col + 2, f"{current_skill[start_col]}: {perk}", "perk_rank", current_skill[start_col], obtained, detail, key=f"{current_skill[start_col]}|{perk}")

    return [entry for entry in entries if entry.tab in CHECKLIST_TABS]


def choose_objective_id(entry: ChecklistEntry, indexes: dict[str, Any]) -> tuple[str, str]:
    normalized_entry = normalize(entry.entry)
    normalized_entry_without_parens = normalize(re.sub(r"\s*\([^)]*\)", "", entry.entry))
    normalized_key = normalize(entry.key)
    normalized_group = normalize(entry.group)

    if entry.category == "quest":
        for key in (normalized_entry, normalized_entry_without_parens):
            if key in QUEST_ENTRY_OVERRIDES:
                return QUEST_ENTRY_OVERRIDES[key], "checklist_manual_objective_alias"
        if normalized_entry in LITANY_OF_LARCENY_ENTRIES:
            return "OBJ-000143", "checklist_manual_objective_alias"
        if normalized_entry == "house in the city" and normalized_group in CITY_HOME_BY_GROUP:
            return CITY_HOME_BY_GROUP[normalized_group], "checklist_manual_property_alias"
        if normalized_entry == "develop land outside the city" and normalized_group in HEARTHFIRE_HOME_BY_GROUP:
            return HEARTHFIRE_HOME_BY_GROUP[normalized_group], "checklist_manual_property_alias"
        if normalized_entry in REPRESENTATIVE_ACTIVITY_OBJECTIVES:
            return REPRESENTATIVE_ACTIVITY_OBJECTIVES[normalized_entry], "checklist_manual_representative_activity"
        location_id = indexes["locations"].get(normalize(entry.detail), "")
        if location_id:
            return location_id, "checklist_detail_location_alias"

    if entry.category == "enchantment":
        objective_id = indexes["enchantments"].get(normalized_entry, "")
        if not objective_id:
            alias = ENCHANTMENT_ENTRY_ALIASES.get(normalized_entry, "")
            if alias:
                objective_id = indexes["enchantments"].get(alias, "")
                if objective_id:
                    return objective_id, "checklist_manual_enchantment_alias"
        return objective_id, "enchantment_catalog" if objective_id else ""

    if entry.category in {"spell", "spell_tome_book"}:
        if normalized_entry in SPELL_OVERRIDES:
            return SPELL_OVERRIDES[normalized_entry], "checklist_manual_spell_parent"
        teleport_pet = re.sub(r"^teleport pet\s+", "", normalized_entry).strip()
        if teleport_pet in PET_TELEPORT_OBJECTIVES:
            return PET_TELEPORT_OBJECTIVES[teleport_pet], "checklist_manual_pet_spell_parent"
        spell_name = entry.entry.removeprefix("Spell Tome: ").strip()
        for candidate in (f"Spell Tome: {spell_name}", spell_name):
            ids = indexes["aliases"].get(normalize(candidate), [])
            if ids:
                return ids[0], "spell_tome_objective"

    if entry.category in {"skill_book", "black_book", "book_document"}:
        for key in (normalized_entry, normalized_entry_without_parens):
            if key in BOOK_TITLE_OVERRIDES:
                return BOOK_TITLE_OVERRIDES[key], "checklist_manual_book_alias"
        objective_id = indexes["book_titles"].get(normalized_entry, "")
        if objective_id:
            return objective_id, "book_location_table"

    if entry.category == "location":
        objective_id = indexes["locations"].get(normalized_entry, "")
        return objective_id, "location_catalog" if objective_id else ""

    if entry.category == "merchant_investment":
        objective_id = indexes["merchants"].get(normalized_entry, "")
        if objective_id:
            return objective_id, "merchant_investment_catalog"
        if normalized_entry in indexes["merchant_excluded"]:
            return "", "merchant_investment_excluded"
        return "", ""

    if entry.category == "collectible_item":
        if normalized_entry in COLLECTIBLE_ENTRY_OVERRIDES:
            return COLLECTIBLE_ENTRY_OVERRIDES[normalized_entry], "checklist_manual_collectible_parent"
        quoted_quests = re.findall(r'"([^"]+)"', entry.detail)
        for quest_name in quoted_quests:
            ids = indexes["aliases"].get(normalize(quest_name), [])
            quest_ids = [
                objective_id
                for objective_id in ids
                if indexes["objective_rows"].get(objective_id, {}).get("category") in {"quest", "misc_objective", "radiant"}
            ]
            if quest_ids:
                return quest_ids[0], "checklist_detail_quest_alias"
        objective_id = indexes["item_members"].get(normalized_entry, "")
        if objective_id:
            return objective_id, "item_member_table"

    if entry.category == "alchemy_effect":
        ingredient, _, effect = entry.key.partition("|")
        normalized_ingredient = normalize(ingredient)
        normalized_effect = normalize(effect)
        ingredient_alias = ALCHEMY_INGREDIENT_ALIASES.get(normalized_ingredient, normalized_ingredient)
        objective_id = indexes["alchemy"].get((normalized_ingredient, normalized_effect), "")
        if not objective_id:
            objective_id = indexes["alchemy"].get((ingredient_alias, normalized_effect), "")
            if objective_id:
                return objective_id, "checklist_manual_alchemy_alias"
        return objective_id, "alchemy_effect_catalog" if objective_id else ""

    if entry.category == "perk_rank":
        skill, _, perk = entry.key.partition("|")
        normalized_skill = normalize(skill)
        normalized_perk = normalize(perk)
        if normalized_skill in TRANSFORMATION_PERK_OBJECTIVES:
            return TRANSFORMATION_PERK_OBJECTIVES[normalized_skill], "checklist_transformation_perk_parent"
        if normalized_skill == "quest perks" and normalized_perk in QUEST_PERK_OBJECTIVES:
            return QUEST_PERK_OBJECTIVES[normalized_perk], "checklist_quest_perk_parent"
        objective_id = indexes["perks"].get((normalized_skill, normalized_perk), "")
        if not objective_id:
            alias = PERK_NAME_ALIASES.get((normalized_skill, normalized_perk), "")
            if alias:
                objective_id = indexes["perks"].get((normalized_skill, normalize(alias)), "")
        return objective_id, "perk_rank_catalog" if objective_id else ""

    if entry.category == "dragon_shout_word":
        ids = indexes["aliases"].get(normalized_key, [])
        if ids:
            return ids[0], "shout_objective"

    if entry.category == "follower_option":
        if indexes["npc_options"].get(normalized_entry):
            return "", "npc_option_table"

    if entry.category == "unique_item" and normalized_entry in UNIQUE_ITEM_OVERRIDES:
        return UNIQUE_ITEM_OVERRIDES[normalized_entry], "checklist_manual_unique_item_alias"

    for candidate in (entry.entry, re.sub(r"\s*\([^)]*\)", "", entry.entry), entry.key):
        ids = indexes["aliases"].get(normalize(candidate), [])
        if ids:
            # Prefer a unique-item row for unique gear and a quest row for quests
            if entry.category == "unique_item":
                unique_ids = [
                    objective_id
                    for objective_id in ids
                    if indexes["objective_rows"].get(objective_id, {}).get("category") == "unique_item"
                ]
                if unique_ids:
                    return unique_ids[0], "objective_alias"
            if entry.category == "quest":
                quest_ids = [
                    objective_id
                    for objective_id in ids
                    if indexes["objective_rows"].get(objective_id, {}).get("category") in {"quest", "misc_objective", "radiant"}
                ]
                if quest_ids:
                    return quest_ids[0], "objective_alias"
            return ids[0], "objective_alias"

    return "", ""


BRANCH_BY_OBJECTIVE = {
    **{f"OBJ-{number:06d}": "BR-001 Stormcloak Civil War" for number in range(87, 102)},
    **{f"OBJ-{number:06d}": "BR-002 Volkihar" for number in [356, 357, *range(374, 384), 1716, 1717, 1736, 1737]},
    "OBJ-000067": "BR-003 Destroy the Dark Brotherhood",
    "OBJ-000019": "BR-004 Paarthurnax / Blades",
    "OBJ-000317": "BR-004 Paarthurnax / Blades",
    "OBJ-002785": "BR-004 Paarthurnax / Blades",
    "OBJ-002786": "BR-004 Paarthurnax / Blades",
    "OBJ-000454": "BR-006 Thirsk Riekling",
    **{f"OBJ-{number:06d}": "BR-007 Ghosts destroy-heretics" for number in range(615, 621)},
    "OBJ-000740": "BR-007 Ghosts destroy-heretics",
    "OBJ-000574": "BR-008A Bittercup Power",
    "OBJ-001612": "BR-009 Azura's Star",
    "OBJ-001555": "BR-010 Rueful Axe",
    "OBJ-001581": "BR-011 Savior's Hide",
    "OBJ-001565": "BR-015 Aetherial Staff",
    "OBJ-001585": "BR-015 Aetherial Shield",
    "OBJ-002777": "BR-005 Master Criminal",
}


def classify(entry: ChecklistEntry, objective_id: str, match_source: str, indexes: dict[str, Any]) -> dict[str, str]:
    objective = indexes["objective_rows"].get(objective_id, {})
    prototype = indexes["prototype_rows"].get(objective_id, {})
    route_placement = objective.get("route_placement", "")
    route_block = prototype.get("route_block", "")
    disposition = prototype.get("disposition", "")
    prototype_status = prototype.get("prototype_status", "")
    deferred_to = prototype.get("deferred_to", "")

    row = {
        "objective_id": objective_id,
        "matched_objective_name": objective.get("objective_name", ""),
        "route_block": route_block,
        "disposition": disposition,
        "prototype_status": prototype_status,
        "deferred_to": deferred_to,
        "match_status": "matched" if objective_id else ("support_table_only" if match_source else "unmatched"),
        "match_source": match_source,
        "mapping_type": "Source-readiness hold",
        "guide_location": "",
        "branch_name": "",
        "exclusion_reason": "",
        "source_note_refs": objective.get("citations", ""),
        "status": "source_readiness_required",
        "notes": "",
    }
    if match_source in {"checklist_manual_enchantment_alias", "checklist_manual_alchemy_alias"}:
        refs = [part.strip() for part in row["source_note_refs"].split("|") if part.strip()]
        if "SN-000126-progression-source-selection-and-grind-policy.md" not in refs:
            refs.append("SN-000126-progression-source-selection-and-grind-policy.md")
        row["source_note_refs"] = " | ".join(refs)

    if objective_id:
        if route_placement == "branch_route":
            row["mapping_type"] = "Branch-route prototype"
            row["branch_name"] = BRANCH_BY_OBJECTIVE.get(objective_id, "TB-029 branch prototype")
            row["guide_location"] = row["branch_name"]
            row["status"] = "mapped_to_branch_prototype"
        elif route_placement == "option_list":
            row["mapping_type"] = "Option-list note"
            row["guide_location"] = deferred_to or disposition or "TB-031D route-default register / TB-035 option presentation"
            row["status"] = "mapped_to_option_list"
        elif route_placement == "appendix":
            row["mapping_type"] = "Appendix-only checklist"
            row["guide_location"] = deferred_to or disposition or "appendix"
            row["status"] = "mapped_to_appendix"
        elif route_placement == "excluded":
            row["mapping_type"] = "Explicit exclusion"
            row["guide_location"] = "excluded"
            row["exclusion_reason"] = objective.get("notes", "") or "Objective row is excluded by current project scope."
            row["status"] = "excluded_with_justification"
        else:
            row["mapping_type"] = "Main-route prototype block"
            row["guide_location"] = route_block or deferred_to or disposition or "route_anchor_or_later_pass"
            row["status"] = "mapped_to_route_prototype"
        row["notes"] = "Mapped to current objective/route-planning prototype; final guide step number remains later work."
        return row

    if entry.category == "general_book":
        row["mapping_type"] = "Explicit exclusion"
        row["guide_location"] = "excluded"
        row["exclusion_reason"] = (
            "TB-031A scope review excludes broad regular-book library collection from the required route. "
            "The required book/document scope remains skill books, spell tomes, Black Books, quest/AE books, "
            "and checklist-tracked unique books."
        )
        row["match_status"] = "support_table_only"
        row["match_source"] = "book_scope_review"
        row["status"] = "excluded_with_justification"
        row["notes"] = (
            "TB-031A reviewed this broad regular-book checklist row and excluded it from required route coverage; "
            "optional broader library completion may be documented separately only if the project scope changes."
        )
        return row

    if entry.category == "merchant_reference":
        row["mapping_type"] = "Appendix-only checklist"
        row["guide_location"] = "TB-031B/TB-036 merchant appendix"
        row["match_status"] = "support_table_only"
        row["match_source"] = "merchant_reference_appendix"
        row["status"] = "mapped_to_appendix"
        row["notes"] = "Non-investable merchant checklist row; route only if NPC/service dependency or warning layer needs it."
        return row

    if entry.category == "merchant_investment" and match_source == "merchant_investment_excluded":
        row["mapping_type"] = "Explicit exclusion"
        row["guide_location"] = "excluded"
        row["exclusion_reason"] = indexes["merchant_excluded"].get(normalize(entry.entry), "Investment unavailable under official PS4 AE scope.")
        row["status"] = "excluded_with_justification"
        row["notes"] = "Checklist marks this merchant investable, but the source-backed investment catalog excludes the investment under the current project scope."
        return row

    if entry.category == "follower_option" and match_source == "npc_option_table":
        row["mapping_type"] = "Option-list note"
        row["guide_location"] = "TB-031D route-default register / TB-035 option presentation"
        row["status"] = "mapped_to_option_list"
        row["notes"] = "Matched to relationship option table; TB-031D records route-affecting defaults and TB-035 presents non-default options."
        return row

    if entry.category in SOURCE_READINESS_BY_CATEGORY:
        guide_location, source, note = SOURCE_READINESS_BY_CATEGORY[entry.category]
        row["mapping_type"] = "Source-readiness hold"
        row["guide_location"] = guide_location
        row["source_note_refs"] = CHECKLIST_MANUAL_REVIEW_SOURCE_NOTE
        row["match_status"] = "support_table_only"
        row["match_source"] = source
        row["status"] = "source_readiness_required"
        row["notes"] = note
        return row

    row["mapping_type"] = "Source-readiness hold"
    row["guide_location"] = "TB-031H source-readiness review"
    row["source_note_refs"] = CHECKLIST_MANUAL_REVIEW_SOURCE_NOTE
    row["match_status"] = "support_table_only"
    row["match_source"] = "checklist_other_source_readiness"
    row["status"] = "source_readiness_required"
    row["notes"] = "Checklist row lacks a current source-backed objective/support-table match; TB-031H must validate, promote, or exclude it explicitly."
    return row


def build_rows(entries: list[ChecklistEntry], indexes: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for sequence, entry in enumerate(entries, start=1):
        objective_id, match_source = choose_objective_id(entry, indexes)
        mapping = classify(entry, objective_id, match_source, indexes)
        checklist_id = f"CHK-{re.sub(r'[^A-Z0-9]+', '-', entry.tab.upper()).strip('-')}-{sequence:04d}"
        row = {
            "checklist_id": checklist_id,
            "checklist_tab": entry.tab,
            "checklist_entry": entry.entry,
            "category": entry.category,
            "raw_sheet_row": str(entry.row),
            "raw_cell": entry.cell,
            "raw_group": entry.group,
            "raw_status": entry.raw_status,
            "raw_detail": entry.detail,
            **mapping,
        }
        rows.append({column: row.get(column, "") for column in COLUMNS})
    return rows


def write_matrix(rows: list[dict[str, str]]) -> None:
    with COVERAGE_MATRIX.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def summarize(rows: list[dict[str, str]]) -> None:
    def counts(field: str) -> dict[str, int]:
        out: dict[str, int] = {}
        for row in rows:
            out[row[field]] = out.get(row[field], 0) + 1
        return dict(sorted(out.items(), key=lambda item: (-item[1], item[0])))

    print(f"Wrote {len(rows)} checklist coverage rows to {COVERAGE_MATRIX.relative_to(REPO_ROOT)}.")
    print("By tab:")
    for key, value in counts("checklist_tab").items():
        print(f"  {key}: {value}")
    print("By mapping_type:")
    for key, value in counts("mapping_type").items():
        print(f"  {key}: {value}")
    print("By status:")
    for key, value in counts("status").items():
        print(f"  {key}: {value}")


def main() -> int:
    if not RAW_WORKBOOK.exists():
        print(f"Missing raw checklist workbook: {RAW_WORKBOOK}", file=sys.stderr)
        return 1
    indexes = load_indexes()
    entries = parse_workbook(RAW_WORKBOOK)
    rows = build_rows(entries, indexes)
    write_matrix(rows)
    summarize(rows)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
